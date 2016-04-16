from openpyxl import load_workbook
import csv
import argparse
import os
import ntpath
import datetime
import numpy as np
import collections

def parse_data():
    wb = load_workbook(filename='ball_by_ball.xlsx', read_only=True)
    ws = wb.worksheets[0] 
    table = np.array([[cell.value for cell in col] for col in ws['B2':'K302']])
    return table

def parse_arguments():
    parser = argparse.ArgumentParser()
    parser.add_argument('-f1', '--folder_innings1',action='store',required='true',
                        help='Path to the first innings data')
    parser.add_argument('-f2', '--folder_innings2',action='store',required='true',
                        help='Path to the second innings data')
    return parser

def compute_wrl(source_folder_path, table,destination_folder_path):
    print len(table)
    for f in os.listdir(source_folder_path):
        file_path = os.path.join(source_folder_path, f)
        print f
        with open(file_path, 'r') as curr:
            curr_reader = csv.reader(curr)
            data1 = []
            row = next(curr_reader)
            row.append('wrl')
            data1.append(row)  
            wickets_lost_prev = 0
            wrl = 0                 
            for r in curr_reader:
                try:
                    wickets_lost = int(r[4])
                #    print "balls "  + row[2] + " wickets " + str(wickets_lost)
                    if(wickets_lost < 10 ):
                        if( wickets_lost != wickets_lost_prev):
                           # print str( r[2]) + " " + str(wickets_lost)
                            wrl =( (table[int(r[2])][0] - table[int(r[2])][wickets_lost])/table[int(r[2])][0])*10                     
                    else:
                        wrl = 10
                    r.append(round(wrl,2))
                    data1.append(r)
                    wickets_lost_prev = wickets_lost
                except:
                    pass

            with open(destination_folder_path + f  ,'w') as newfile:
                writer = csv.writer(newfile)
                writer.writerows(data1)

def main():
    parser = parse_arguments()
    args = parser.parse_args()
    folder_path1 = args.folder_innings1
    folder_path2 = args.folder_innings2
    table = []
    table = parse_data()
    compute_wrl(folder_path1, table, 'CSV_Data/Inn1/') 
    compute_wrl(folder_path2, table, 'CSV_Data/Inn2/') 

if __name__ == "__main__":
    main()
